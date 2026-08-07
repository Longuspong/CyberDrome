#!/usr/bin/env python3
"""
export_parts_table.py -- Teile-Uebersicht als Excel-Mappe.

    python3 tools/export_parts_table.py [ziel.xlsx]

Erzeugt aus dem Inhalt von ``parts/`` eine lesbare Arbeitsmappe fuers Game
Design: welche Teile es gibt, welche Spielwerte sie mitbringen, was an welchen
Slot passt -- und wie sich zwei Teile desselben Typs vergleichen lassen.

Warum als Skript und nicht als gepflegte Datei
----------------------------------------------
Eine von Hand gefuehrte Teileliste laeuft nach dem dritten Set auseinander --
dasselbe Argument, aus dem der Generator die Silhouetten und die Slot-Matrix
selbst ausrechnet, statt sie irgendwo abzuschreiben. Alles, was in dieser
Mappe steht, kommt aus den Teil-JSONs. Wer eine Zahl aendern will, aendert sie
dort und laesst die Mappe neu erzeugen.

Aufbau der Mappe
----------------
Ein Blatt je Bauteiltyp -- Koerperteile, Kopfteile, Fussteile, Kerne -- und
zwei fuer die Ausruestung, getrennt nach Waffe und Sonstigem (Schild und
Support). Jedes dieser Blaetter zeigt NUR die Werte, die sein Typ ueberhaupt
fuehrt: ein Kopf hat keine Traglast, ein Kern keine Bewegung. Das ist der
ganze Zweck der Trennung -- vier Spalten, die vergleichbar sind, statt
achtzehn, von denen sechzehn null sind.

Darueber liegt das Blatt „Alle Teile“ mit dem vollen Wertesatz aller Teile in
einer Tabelle. Es ist die Fassung zum Filtern und Sortieren; die Typblaetter
sind die Fassung zum Lesen.

Der Threat-Beitrag
------------------
Die Typblaetter tragen eine gemeinsame Vergleichszahl: den Beitrag des Teils
zum Threat Score aus ``scripts/core/drome_build.gd``. Das ist die einzige
Gewichtung, die das Spiel selbst benutzt (das Gegner-Matching des Chaos-Virus
haengt daran), und sie ist linear -- der Score eines Aufbaus ist damit die
Summe der Beitraege seiner Teile. Deshalb laesst sich ueber diese eine Spalte
ein Kopf mit einem Antrieb vergleichen, obwohl die beiden keinen einzigen
Wert gemeinsam haben.

Eine Einschraenkung gehoert dazu: Waffen gehen zusaetzlich mit ``power * 2``
in den Score ein, aber nur die STAERKSTE Waffe eines Aufbaus. Das ist keine
Eigenschaft des Teils, sondern der Bestueckung, und steht deshalb nicht in
der Spalte -- auf dem Waffenblatt aber als eigene Spalte daneben.
"""

from __future__ import annotations

import json
import pathlib
import re
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
import build_sample_parts as gen  # noqa: E402  -- fuer den Silhouetten-Abstand

ROOT = pathlib.Path(__file__).resolve().parent.parent
PARTS_DIR = ROOT / "parts"

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3B4D")
SUB_FILL = PatternFill("solid", fgColor="DCE6F1")
CALC_FILL = PatternFill("solid", fgColor="EAF1F8")   # Formel, nicht aus der JSON
HELPER_FILL = PatternFill("solid", fgColor="F2F2F2")
SUM_FILL = PatternFill("solid", fgColor="F7F7F7")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
SUB_FONT = Font(name=FONT, bold=True, size=10)
CODE_FONT = Font(name=FONT, bold=True, size=10)
BODY_FONT = Font(name=FONT, size=10)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
CALC_FONT = Font(name=FONT, size=10, color="1F3B4D")
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

TYPE_LABEL = {"core": "Kern", "head": "Kopf", "body": "Koerper",
              "feet": "Fuesse", "equipment": "Ausruestung"}
TYPE_ORDER = ["body", "head", "feet", "core", "equipment"]
CLASS_LABEL = {"light": "leicht", "medium": "mittel", "heavy": "schwer"}
CATEGORY_LABEL = {"weapon": "Waffe", "shield": "Schild", "support": "Support"}
ACTION_CATEGORY_LABEL = {"attack": "Angriff", "ability": "Faehigkeit"}
TARGETING_LABEL = {
    "single": "Einzelziel",
    "self": "selbst",
    "tile": "Feld",
    "aoe_around_target": "Flaeche um das Ziel",
}

# Rolle des Bauteiltyps im Spiel -- aus docs/GAME_DESIGN.md, Abschnitt 3.
TYPE_ROLE = {
    "core": "Identitaet der Einheit: Energiespeicher und Energieabgabe",
    "head": "Sensorik: Sicht, Zielwert, Statusresistenz",
    "body": "Chassis: HP, Panzerung, Traglast -- und die Zahl der Ausruestungsslots",
    "feet": "Bewegungsreichweite, Gelaendeverhalten, Tempo",
    "equipment": "Waffen, Schilde, Support-Module -- alles, was Aktionen gewaehrt",
}

# Gewichtung des Threat Score aus scripts/core/drome_build.gd. Wer sie dort
# aendert, aendert sie hier mit -- sonst rechnet die Mappe eine zweite,
# stillschweigend andere Staerke aus als das Spiel.
THREAT_WEIGHTS = {"hp": 1.0, "atk": 6.0, "def": 8.0,
                  "spd": 3.0, "mov": 4.0, "en_max": 0.5}
THREAT_BEST_WEAPON = 2.0

# Was ein Spielwert bedeutet und woran er haengt. Wird zur Legende und zu den
# Kommentaren an den Spaltenkoepfen.
GLOSSARY = {
    "hp": ("HP", "Strukturpunkte. Ein DROME hat keine Grundwerte -- seine "
                 "Lebensleiste ist die Summe aus Chassis und Kopf."),
    "en_max": ("EN max", "Energiespeicher. Bezahlt die EN-Kosten der Aktionen."),
    "en_regen": ("EN/Zyklus", "Energie, die je Zyklus nachlaeuft."),
    "spd": ("SPD", "Tempo auf der TICK-Leiste: so viel fuellt der DROME je "
                   "Tick, bei 100 ist er am Zug. Muss im fertigen Aufbau "
                   "mindestens 1 sein."),
    "mov": ("MOV", "Felder je Zug. Muss im fertigen Aufbau mindestens 1 sein."),
    "atk": ("ATK", "Angriffsbonus, additiv auf die Power der Aktion."),
    "def": ("DEF", "Schadensminderung."),
    "weight": ("Gewicht", "Zaehlt gegen die Traglast des Chassis."),
    "power_draw": ("Energiebedarf", "Zaehlt gegen die Energieabgabe des Kerns."),
    "weight_capacity": ("Traglast", "Nur Chassis: wie viel Gewicht der Aufbau "
                                    "insgesamt tragen darf."),
    "power_output": ("Energieabgabe", "Nur Kern: wie viel Energiebedarf der "
                                      "Aufbau insgesamt haben darf."),
    "can_pass_units": ("durch Einheiten", "Darf Felder anderer Einheiten "
                                          "durchqueren."),
    "can_enter_steps": ("Stufen", "Darf Stufen-Felder betreten."),
    "ignores_drift": ("Drift egal", "Drift-Zonen schieben diesen DROME nicht."),
    "drift_modifier": ("Drift +/-", "Zusaetzliche Felder, die eine Drift-Zone "
                                    "schiebt. Positiv heisst: rutscht weiter."),
    "step_cost_reduced": ("Stufen billiger", "Stufen kosten weniger Bewegung."),
    "grants_ignore_haze": ("sieht durch Haze", "Haze-Felder blockieren die "
                                               "Sichtlinie dieses DROME nicht."),
}

# Aktionsfelder -- aus scripts/core/action_data.gd.
ACTION_GLOSSARY = {
    "range_tiles": ("Reichweite", "Felder bis zum Ziel."),
    "aoe_radius": ("AoE", "Radius um das Ziel. 0 = nur das Ziel selbst."),
    "en_cost": ("EN-Kosten", "Energie je Einsatz."),
    "power": ("Power", "Positiv = Schaden, negativ = Heilung. Ein Vorzeichen "
                       "statt zweier Felder."),
    "requires_line_of_sight": ("Sichtlinie", "Braucht freie Sicht zum Ziel."),
    "push_tiles": ("Schub", "Positiv = wegstossen, negativ = heranziehen."),
}

# Silhouetten-Merkmal und was es dem Spieler sagen soll.
# Spiegelt parts/README.md, Abschnitt 6a -- dort steht die Regel, hier nur die
# Kurzfassung je Teil, damit die Mappe ohne das Dokument lesbar bleibt.
READING = {
    "CHS-001": ("schlank, kantig, zweibeinig", "leichter Standardrahmen, zwei Arme"),
    "CHS-002": ("geduckt: Panzerschuerze, Keil-Torso, Pauldrons ueber dem "
                "Kopfsockel, Auspuffstapel",
                "schwerer Rahmen, drei Anker -- Schulter nur fuer Support"),
    "CHS-003": ("rund, Fahrgestell statt Beinen", "Support-Rahmen, weiche Formensprache"),
    "CHS-004": ("schmal und hoch, Gegengewichts-Ausleger ueber dem Kopf",
                "Ein-Slot-Rahmen, Scharfschuetze"),
    "HED-001": ("kleiner Helm mit Visierband und Antenne", "Aufklaerung"),
    "HED-002": ("wuchtiger Bunkerkopf mit Seitensensoren", "gepanzert, robust"),
    "HED-003": ("flacher Helm mit weit auskragender Krempe", "Rundumsicht, arkan"),
    "HED-004": ("ein langes Okular statt Visierband, Rueckenfinne",
                "sieht WEIT statt viel"),
    "LEG-001": ("schmaler Stand, gerade Stelzen, eine Zehe", "schnell"),
    "LEG-002": ("breiter Stand auf Auslegern, Hydraulikzylinder, Ferse",
                "langsam, standfest, stuetzt Rueckstoss ab"),
    "LEG-003": ("zwei grosse Raeder auf einer Achse", "rollend, schnell auf Ebene"),
    "LEG-004": ("Knick nach hinten, digitigrad", "federnd, mittlere Reichweite"),
    "COR-001": ("Scheibe mit vier Energiestrichen", "Impuls"),
    "COR-002": ("grosse Scheibe mit Kreuzglut", "Fusion, hohe Leistung"),
    "COR-003": ("Orb mit umlaufenden Ringen", "arkan"),
    "COR-004": ("senkrechter Leuchtspalt statt Scheibe", "Zielrechner, Optik"),
    "EQP-001": ("ein glattes Rohr, sonst nichts", "leicht, kurze Reichweite"),
    "EQP-002": ("flache Wand quer zur Blickrichtung", "Deckung"),
    "EQP-003": ("Muendungsbremse mit Querfluegeln, Trommelmagazin, Abstuetzstrebe",
                "schwer, abgestuetzt -- Artillerie"),
    "EQP-004": ("flacher Pod auf der Schulterbruecke", "Support, keine Waffe"),
    "EQP-005": ("langer duenner Stab mit Kopfglut", "arkane Waffe"),
    "EQP-006": ("schwebender Ring vor der Hand", "Support, arkan"),
    "EQP-007": ("ueberlanger duenner Doppellauf, Zielblock, Gabel",
                "Praezision, grosse Reichweite"),
}


# ---------------------------------------------------------------------------
# Daten einlesen
# ---------------------------------------------------------------------------
def load_parts() -> tuple[list[dict], dict]:
    """Ein Eintrag je Teil (Suedansicht als Vertreter) plus Set-Namen."""
    parts, set_names = [], {}
    for set_dir in sorted(p for p in PARTS_DIR.iterdir() if p.is_dir()):
        set_json = set_dir / "set.json"
        set_names[set_dir.name] = (
            json.loads(set_json.read_text(encoding="utf-8")).get("name", set_dir.name)
            if set_json.is_file() else set_dir.name
        )
        for json_path in sorted(set_dir.glob("*.json")):
            if json_path.name == "set.json":
                continue
            meta = json.loads(json_path.read_text(encoding="utf-8"))
            if meta.get("direction") != "south":
                continue          # vier Richtungen, ein Teil
            parts.append(meta)
    parts.sort(key=lambda m: (TYPE_ORDER.index(base_type(m["type"])), m.get("code", "")))
    return parts, set_names


def base_type(part_type: str) -> str:
    return "equipment" if part_type.startswith("equipment") else part_type


def equip_slots(part: dict) -> list[str]:
    return sorted(a["name"] for a in part.get("anchors", [])
                  if a["name"].startswith("equip_"))


def stat(part: dict, key: str):
    return part.get("stats", {}).get(key, 0)


def action(part: dict) -> dict | None:
    return part.get("stats", {}).get("action")


def flag(part: dict, key: str) -> str:
    return "ja" if stat(part, key) else "–"


def threat_of(part: dict, keys) -> float:
    """Beitrag eines Teils zum Threat Score -- nur ueber die genannten Werte."""
    return sum(THREAT_WEIGHTS[key] * stat(part, key) for key in keys)


# ---------------------------------------------------------------------------
# Bausteine
# ---------------------------------------------------------------------------
def title(sheet, row: int, text: str, note: str = "") -> int:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT, bold=True, size=14)
    row += 1
    if note:
        cell = sheet.cell(row=row, column=1, value=note)
        cell.font = NOTE_FONT
        row += 1
    return row + 1


def header_row(sheet, row: int, labels: list[str], start: int = 1) -> int:
    for offset, label in enumerate(labels):
        cell = sheet.cell(row=row, column=start + offset, value=label)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    sheet.row_dimensions[row].height = 30
    return row + 1


def widths(sheet, mapping: dict) -> None:
    for column, width in mapping.items():
        sheet.column_dimensions[column].width = width


# ---------------------------------------------------------------------------
# Spaltenbeschreibung fuer die Teilblaetter
# ---------------------------------------------------------------------------
class Column:
    """
    Eine Spalte eines Teilblatts.

    ``get`` liefert den Wert aus dem Teil -- das ist der FAKT aus der JSON.
    ``formula`` liefert stattdessen eine Formel und bekommt dafuer die
    Spaltenbuchstaben des Blatts, damit ein verschobenes Layout die Bezuege
    mitnimmt statt sie zu brechen.
    """

    def __init__(self, key, header, width, get=None, *, formula=None,
                 fmt=None, summary=False, wrap=False, note="", bar=False):
        self.key = key
        self.header = header
        self.width = width
        self.get = get
        self.formula = formula
        self.fmt = fmt
        self.summary = summary      # bekommt Min/Ø/Max unter der Tabelle
        self.wrap = wrap
        self.note = note
        self.bar = bar              # Datenbalken zum Vergleichen


def stat_column(key, width=9, summary=True, **kwargs) -> Column:
    label, note = GLOSSARY[key]
    return Column(key, label, width, lambda p, k=key: stat(p, k),
                  fmt="0", summary=summary, note=note, **kwargs)


def flag_column(key, width=11) -> Column:
    label, note = GLOSSARY[key]
    return Column(key, label, width, lambda p, k=key: flag(p, k), note=note)


def action_column(key, width=10, summary=True, label=None) -> Column:
    default, note = ACTION_GLOSSARY[key]
    # Ein passives Teil hat keine Aktion. Die Zelle bleibt dann leer statt 0 --
    # eine 0 hiesse „Reichweite null“, und das ist etwas anderes als „gibt es
    # nicht“.
    return Column("act_" + key, label or default, width,
                  lambda p, k=key: (action(p) or {}).get(k),
                  fmt="0", summary=summary, note=note)


def reading_columns() -> list[Column]:
    return [
        Column("silhouette", "Silhouetten-Merkmal", 40,
               lambda p: READING.get(p.get("code", ""), ("", ""))[0], wrap=True,
               note="Woran der Spieler das Teil auf dem Feld erkennt. "
                    "parts/README.md, Abschnitt 6a."),
        Column("reading", "sagt dem Spieler", 30,
               lambda p: READING.get(p.get("code", ""), ("", ""))[1], wrap=True),
        Column("tags", "Tags", 24, lambda p: ", ".join(p.get("tags", []))),
    ]


def head_columns(set_names) -> list[Column]:
    return [
        Column("code", "Code", 10, lambda p: p.get("code", ""),
               note="Stabile Kennung fuers Game Design."),
        Column("name", "Name", 22, lambda p: p.get("name", p["id"])),
        Column("set", "Set", 24, lambda p: set_names.get(p["set"], p["set"])),
        Column("id", "Datei-Slug", 16, lambda p: p["id"],
               note="Der Dateiname ohne Richtung -- so heisst das Teil im Code."),
    ]


def threat_column(keys) -> Column:
    terms = " · ".join(f"{THREAT_WEIGHTS[k]:g}×{GLOSSARY[k][0]}" for k in keys)

    def build(letters, row, keys=keys):
        return "=" + "+".join(
            f"{letters[key]}{row}*{THREAT_WEIGHTS[key]:g}" for key in keys)

    return Column("threat", "Threat-Beitrag", 13, formula=build, fmt="0.0",
                  summary=True, bar=True,
                  note="Beitrag zum Threat Score aus drome_build.gd: "
                       f"{terms}. Die Werte, die dieser Bauteiltyp gar nicht "
                       "fuehrt, sind ueberall 0 und stehen deshalb nicht in "
                       "der Formel. Ueber diese Spalte lassen sich Teile "
                       "verschiedener Typen vergleichen.")


# ---------------------------------------------------------------------------
# Ein Teilblatt
# ---------------------------------------------------------------------------
def sheet_type(book, name, heading, note, parts, columns) -> None:
    sheet = book.create_sheet(name)
    letters = {column.key: get_column_letter(index)
               for index, column in enumerate(columns, start=1)}
    widths(sheet, {letters[c.key]: c.width for c in columns})

    row = title(sheet, 1, heading, note)
    header = row
    row = header_row(sheet, row, [c.header for c in columns])
    for column in columns:
        if column.note:
            cell = sheet.cell(row=header, column=columns.index(column) + 1)
            cell.comment = Comment(f"{column.header}\n\n{column.note}", "CyberDrome")
    first = row

    for part in parts:
        for index, column in enumerate(columns, start=1):
            if column.formula is not None:
                cell = sheet.cell(row=row, column=index,
                                  value=column.formula(letters, row))
                cell.font = CALC_FONT
                cell.fill = CALC_FILL
            else:
                cell = sheet.cell(row=row, column=index, value=column.get(part))
                cell.font = CODE_FONT if column.key == "code" else BODY_FONT
            cell.border = BOX
            if column.fmt:
                cell.number_format = column.fmt
                cell.alignment = CENTER
            elif column.wrap:
                cell.alignment = WRAP
            elif column.key in ("silhouette", "reading"):
                cell.alignment = WRAP
        sheet.row_dimensions[row].height = 30
        row += 1
    last = row - 1

    # Min / Ø / Max je Wertespalte. Formeln, keine ausgerechneten Zahlen --
    # wer eine Zeile ergaenzt, will die Kennzahl nicht nachrechnen muessen.
    row += 1
    summary_first = row
    for label, function in (("Minimum", "MIN"), ("Mittel", "AVERAGE"),
                            ("Maximum", "MAX")):
        cell = sheet.cell(row=row, column=1, value=label)
        cell.font = SUB_FONT
        cell.fill = SUM_FILL
        for index, column in enumerate(columns, start=1):
            if not column.summary:
                continue
            letter = letters[column.key]
            cell = sheet.cell(row=row, column=index,
                              value=f"={function}({letter}{first}:{letter}{last})")
            cell.font = SUB_FONT
            cell.fill = SUM_FILL
            cell.number_format = "0.0" if function == "AVERAGE" else column.fmt
            cell.alignment = CENTER
            cell.border = BOX
        row += 1
    sheet.cell(row=summary_first, column=2,
               value="Kennzahlen ueber die Zeilen darueber -- Formeln, "
                     "keine festen Zahlen").font = NOTE_FONT

    for column in columns:
        if not column.bar:
            continue
        letter = letters[column.key]
        sheet.conditional_formatting.add(
            f"{letter}{first}:{letter}{last}",
            DataBarRule(start_type="min", end_type="max", color="638EC6"))

    sheet.freeze_panes = sheet.cell(row=first, column=3).coordinate
    sheet.auto_filter.ref = f"A{header}:{get_column_letter(len(columns))}{last}"


# ---------------------------------------------------------------------------
# Blatt: Alle Teile
# ---------------------------------------------------------------------------
def sheet_all(book, parts, set_names) -> tuple[int, int]:
    """Volle Tabelle zum Filtern. Gibt die Zeilenspanne der Daten zurueck."""
    columns = head_columns(set_names)[:3] + [
        Column("type", "Typ", 13, lambda p: TYPE_LABEL[base_type(p["type"])]),
        Column("category", "Bauart", 10,
               lambda p: CATEGORY_LABEL.get(p.get("category"), "")
               if base_type(p["type"]) == "equipment" else ""),
        Column("mount_class", "Montageklasse", 12,
               lambda p: CLASS_LABEL.get(p.get("mount_class"), "")
               if base_type(p["type"]) == "equipment" else "",
               note="Was die HALTERUNG aushalten muss -- Masse, Rueckstoss, "
                    "Hebel. Nicht das Gewicht allein."),
        stat_column("hp", summary=False),
        stat_column("en_max", summary=False),
        stat_column("en_regen", summary=False),
        stat_column("spd", summary=False),
        stat_column("mov", summary=False),
        stat_column("atk", summary=False),
        stat_column("def", summary=False),
        stat_column("weight", summary=False),
        stat_column("power_draw", 12, summary=False),
        stat_column("weight_capacity", 10, summary=False),
        stat_column("power_output", 12, summary=False),
        Column("act_name", "Aktion", 20,
               lambda p: (action(p) or {}).get("display_name", "")),
        action_column("power", summary=False),
        action_column("range_tiles", summary=False),
        action_column("en_cost", summary=False),
        # Hier stehen alle sechs Werte der Gewichtung als Spalte, also geht
        # die volle Formel. Das ist die einzige Stelle der Mappe, an der sich
        # ein Kopf mit einer Waffe vergleichen laesst.
        threat_column(list(THREAT_WEIGHTS)),
        Column("tags", "Tags", 24, lambda p: ", ".join(p.get("tags", []))),
    ]

    sheet = book.create_sheet("Alle Teile")
    letters = {column.key: get_column_letter(index)
               for index, column in enumerate(columns, start=1)}
    widths(sheet, {letters[c.key]: c.width for c in columns})

    row = title(sheet, 1, "Alle Teile auf einen Blick",
                "Der volle Wertesatz in einer Tabelle -- zum Filtern und "
                "Sortieren. Zum Lesen und Vergleichen sind die Typblaetter "
                "gedacht: dort steht nur, was der jeweilige Typ auch fuehrt.")
    header = row
    row = header_row(sheet, row, [c.header for c in columns])
    for index, column in enumerate(columns, start=1):
        if column.note:
            sheet.cell(row=header, column=index).comment = \
                Comment(f"{column.header}\n\n{column.note}", "CyberDrome")
    first = row

    for part in parts:
        for index, column in enumerate(columns, start=1):
            if column.formula is not None:
                cell = sheet.cell(row=row, column=index,
                                  value=column.formula(letters, row))
                cell.font = CALC_FONT
                cell.fill = CALC_FILL
            else:
                value = column.get(part)
                # Nullen ueberall dort weglassen, wo der Typ den Wert gar
                # nicht fuehrt -- eine Tabelle voller Nullen liest sich wie
                # Rauschen, und man sieht die eine Zahl nicht mehr, auf die
                # es ankommt. Fuer die Formeln macht das keinen Unterschied:
                # eine leere Zelle rechnet Excel als 0.
                if column.fmt == "0" and value == 0:
                    value = None
                cell = sheet.cell(row=row, column=index, value=value)
                cell.font = CODE_FONT if column.key == "code" else BODY_FONT
            cell.border = BOX
            if column.fmt:
                cell.number_format = column.fmt
                cell.alignment = CENTER
        row += 1
    last = row - 1

    for column in columns:
        if column.bar:
            letter = letters[column.key]
            sheet.conditional_formatting.add(
                f"{letter}{first}:{letter}{last}",
                DataBarRule(start_type="min", end_type="max", color="638EC6"))

    sheet.freeze_panes = sheet.cell(row=first, column=3).coordinate
    sheet.auto_filter.ref = f"A{header}:{get_column_letter(len(columns))}{last}"
    return first, last


# ---------------------------------------------------------------------------
# Blatt: Legende
# ---------------------------------------------------------------------------
def sheet_legend(book, parts_range, type_letter) -> None:
    sheet = book.create_sheet("Legende")
    widths(sheet, {"A": 24, "B": 20, "C": 74})
    row = title(sheet, 1, "CyberDrome -- Teile-Uebersicht",
                "Erzeugt aus parts/ durch tools/export_parts_table.py. "
                "Nicht von Hand pflegen, sondern neu erzeugen.")

    entries = [
        ("Alle Teile", "Volle Tabelle aller Teile zum Filtern und Sortieren."),
        ("Koerperteile", "Chassis: HP, Panzerung, Traglast, Ausruestungsslots."),
        ("Kopfteile", "Sensorik: Zielwert, Sicht, Panzerung am Kopf."),
        ("Fussteile", "Antriebe: Bewegung, Tempo, Gelaendeverhalten."),
        ("Kerne", "Energiespeicher und Energieabgabe -- das Budget des Aufbaus."),
        ("Ausruestung Waffen", "Alles mit Bauart „Waffe“, samt Aktionswerten."),
        ("Ausruestung Sonstiges", "Schilde und Support-Module."),
        ("Chassis & Slots", "Je Chassis: wie viele Ausruestungsanker, "
                            "und was jeder annimmt."),
        ("Kompatibilitaet", "Lebende Matrix: welches Ausruestungsteil an "
                            "welchen Slot passt. Rechnet aus „Klasse“ und "
                            "„Bauart“ -- aendert man die, aendert sich die "
                            "Matrix mit."),
        ("Silhouetten", "Wie aehnlich sich zwei Teile desselben Typs sehen. "
                        "1.00 waere ununterscheidbar; ab 0.80 bricht der "
                        "Generator bei Ausruestung ab."),
    ]
    row = header_row(sheet, row, ["Blatt", "", "Was drin steht"])
    for name, text in entries:
        sheet.cell(row=row, column=1, value=name).font = SUB_FONT
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell = sheet.cell(row=row, column=2, value=text)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        sheet.row_dimensions[row].height = 28
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Bestand").font = Font(name=FONT, bold=True, size=12)
    sheet.cell(row=row, column=2,
               value="zaehlt das Blatt „Alle Teile“ -- kommt ein Teil dazu, "
                     "stimmt die Zahl von selbst").font = NOTE_FONT
    row += 1
    row = header_row(sheet, row, ["Bauteiltyp", "Anzahl", "Rolle im Spiel"])
    first = row
    for part_type in TYPE_ORDER:
        sheet.cell(row=row, column=1, value=TYPE_LABEL[part_type]).font = BODY_FONT
        count = sheet.cell(
            row=row, column=2,
            value=f"=COUNTIF('Alle Teile'!${type_letter}${parts_range[0]}:"
                  f"${type_letter}${parts_range[1]},A{row})")
        count.font = BODY_FONT
        count.alignment = CENTER
        cell = sheet.cell(row=row, column=3, value=TYPE_ROLE[part_type])
        cell.font = BODY_FONT
        cell.alignment = WRAP
        for column in (1, 2, 3):
            sheet.cell(row=row, column=column).border = BOX
        row += 1
    sheet.cell(row=row, column=1, value="Summe").font = SUB_FONT
    total = sheet.cell(row=row, column=2, value=f"=SUM(B{first}:B{row - 1})")
    total.font = SUB_FONT
    total.alignment = CENTER
    row += 2

    sheet.cell(row=row, column=1, value="Spielwerte").font = Font(name=FONT, bold=True, size=12)
    sheet.cell(row=row, column=2,
               value="Ein DROME hat keine Grundwerte: jeder Wert ist die Summe "
                     "seiner Teile (drome_build.gd)").font = NOTE_FONT
    row += 1
    row = header_row(sheet, row, ["Feld in der JSON", "Spalte", "Bedeutung"])
    for key, (label, text) in GLOSSARY.items():
        sheet.cell(row=row, column=1, value=key).font = Font(name=FONT, size=9)
        sheet.cell(row=row, column=2, value=label).font = SUB_FONT
        cell = sheet.cell(row=row, column=3, value=text)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        for column in (1, 2, 3):
            sheet.cell(row=row, column=column).border = BOX
        sheet.row_dimensions[row].height = 26
        row += 1
    row += 1

    sheet.cell(row=row, column=1, value="Aktionswerte").font = Font(name=FONT, bold=True, size=12)
    sheet.cell(row=row, column=2,
               value="Schaden ist deterministisch -- kein Trefferwurf, kein "
                     "Krit, keine Ausweichchance (action_data.gd)").font = NOTE_FONT
    row += 1
    row = header_row(sheet, row, ["Feld in der JSON", "Spalte", "Bedeutung"])
    for key, (label, text) in ACTION_GLOSSARY.items():
        sheet.cell(row=row, column=1, value=key).font = Font(name=FONT, size=9)
        sheet.cell(row=row, column=2, value=label).font = SUB_FONT
        cell = sheet.cell(row=row, column=3, value=text)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        for column in (1, 2, 3):
            sheet.cell(row=row, column=column).border = BOX
        row += 1
    row += 1

    sheet.cell(row=row, column=1, value="Bauregeln").font = Font(name=FONT, bold=True, size=12)
    sheet.cell(row=row, column=2,
               value="Was die Werkstatt prueft, bevor ein Aufbau gueltig ist "
                     "-- DromeBuild.validate()").font = NOTE_FONT
    row += 1
    for text in [
        "Chassis, Kopf, Antrieb und Kern sind Pflicht. Ohne eines davon ist "
        "der Aufbau ungueltig.",
        "Mindestens eine Waffe muss bestueckt sein.",
        "Summe Gewicht darf die Traglast des Chassis nicht ueberschreiten.",
        "Summe Energiebedarf darf die Energieabgabe des Kerns nicht "
        "ueberschreiten.",
        "MOV und SPD muessen im fertigen Aufbau mindestens 1 sein -- das "
        "Molok-Chassis zieht beide ins Minus und braucht deshalb einen "
        "Antrieb, der das auffaengt.",
        "Ein Ausruestungsteil muss die Regel seines Slots erfuellen "
        "(Montageklasse und Bauart). Siehe Blatt „Kompatibilitaet“.",
    ]:
        cell = sheet.cell(row=row, column=1, value="• " + text)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        sheet.row_dimensions[row].height = 26
        row += 1
    row += 1

    sheet.cell(row=row, column=1,
               value="Threat-Beitrag").font = Font(name=FONT, bold=True, size=12)
    sheet.cell(row=row, column=2,
               value="die einzige Spalte, die Teile verschiedener Typen "
                     "vergleichbar macht").font = NOTE_FONT
    row += 1
    weights = " + ".join(f"{weight:g} × {GLOSSARY[key][0]}"
                         for key, weight in THREAT_WEIGHTS.items())
    for text in [
        f"Gewichtung aus DromeBuild.threat_score(): {weights}. Das Spiel "
        "benutzt genau diese Zahl, um dem Spielersquad ein gleich starkes "
        "Gegnersquad gegenueberzustellen (Chaos-Virus).",
        "Sie ist linear -- der Score eines Aufbaus ist die Summe der "
        "Beitraege seiner Teile. Deshalb laesst sich ueber diese eine Spalte "
        "ein Kopf mit einem Antrieb vergleichen, obwohl die beiden keinen "
        "einzigen Wert gemeinsam haben.",
        "Auf den Typblaettern stehen in der Formel nur die Werte, die der "
        "Typ ueberhaupt fuehrt. Die uebrigen sind dort bei jedem Teil null, "
        "das Ergebnis ist also dasselbe -- die Selbstpruefung des Skripts "
        "besteht darauf.",
        f"Waffen gehen zusaetzlich mit {THREAT_BEST_WEAPON:g} × Power ein, "
        "aber nur die STAERKSTE Waffe eines Aufbaus. Das haengt an der "
        "Bestueckung, nicht am Teil, und steht deshalb auf dem Waffenblatt "
        "in einer eigenen Spalte daneben.",
        "Was der Score NICHT abbildet: Reichweite, Traglast, Energiebedarf "
        "und alles Gelaende. Ein Antrieb, der Stufen betreten darf, ist "
        "dadurch keinen Punkt wert -- im Gefecht aber oft entscheidend.",
    ]:
        cell = sheet.cell(row=row, column=1, value="• " + text)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        sheet.row_dimensions[row].height = 26
        row += 1
    row += 1

    sheet.cell(row=row, column=1, value="Farben").font = Font(name=FONT, bold=True, size=12)
    row += 1
    for fill, font, text in [
        (None, BODY_FONT, "schwarz = Fakt aus der Teil-JSON"),
        (CALC_FILL, CALC_FONT, "blau hinterlegt = Formel, rechnet aus den "
                               "Spalten daneben"),
        (SUM_FILL, SUB_FONT, "grau = Kennzahl ueber die ganze Spalte"),
        (HELPER_FILL, NOTE_FONT, "hellgrau = Hilfszeile, die die Matrix braucht"),
    ]:
        cell = sheet.cell(row=row, column=1, value="Beispiel")
        if fill:
            cell.fill = fill
        cell.font = font
        cell.border = BOX
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.cell(row=row, column=2, value=text).font = BODY_FONT
        row += 1


# ---------------------------------------------------------------------------
# Blatt: Chassis & Slots
# ---------------------------------------------------------------------------
def rule_text(rule: dict | None) -> str:
    if not rule:
        return "alles"
    parts_ = []
    if rule.get("max_class"):
        parts_.append("bis " + CLASS_LABEL[rule["max_class"]])
    if rule.get("categories"):
        parts_.append("nur " + "/".join(CATEGORY_LABEL[c] for c in rule["categories"]))
    return " · ".join(parts_) or "alles"


def sheet_chassis(book, parts, set_names) -> None:
    sheet = book.create_sheet("Chassis & Slots")
    widths(sheet, {"A": 10, "B": 22, "C": 24, "D": 9, "E": 16, "F": 22, "G": 44})
    row = title(sheet, 1, "Chassis und ihre Ausruestungsslots",
                "Die Zahl der Slots faellt aus den equip_*-Ankern des Chassis, "
                "die Regel je Slot aus seinem Feld slot_rules.")

    row = header_row(sheet, row, [
        "Code", "Chassis", "Set", "Slots", "Slot", "nimmt an", "Anmerkung",
    ])

    notes = {
        ("CHS-002", "equip_shoulder"):
            "Bruecke hinter dem Kopf: kein Gegenhalt, keine Hand. Sensorik und "
            "Drohnen ja, Schild oder schweres Geraet nein.",
        ("CHS-004", "equip_center"):
            "Der einzige Slot -- dafuer ohne Obergrenze. Ein Slot weniger als "
            "der Standard, dafuer die schwerste Waffe.",
        ("CHS-001", "equip_left"): "Sprinter: keine Artillerie an einem Arm, "
                                   "der sechs Felder weit laufen soll.",
    }

    for part in parts:
        if base_type(part["type"]) != "body":
            continue
        slots = equip_slots(part)
        rules = part.get("slot_rules", {})
        start = row
        for index, slot in enumerate(slots):
            values = [
                part.get("code", "") if index == 0 else "",
                part.get("name", part["id"]) if index == 0 else "",
                set_names.get(part["set"], part["set"]) if index == 0 else "",
                len(slots) if index == 0 else "",
                slot,
                rule_text(rules.get(slot)),
                notes.get((part.get("code", ""), slot), ""),
            ]
            for offset, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=offset, value=value)
                cell.font = BODY_FONT
                cell.border = BOX
                cell.alignment = Alignment(wrap_text=offset == 7, vertical="top")
            sheet.cell(row=row, column=1).font = CODE_FONT
            sheet.cell(row=row, column=6).font = SUB_FONT
            sheet.row_dimensions[row].height = 28
            row += 1
        for column in (1, 2, 3, 4):
            if row - start > 1:
                sheet.merge_cells(start_row=start, start_column=column,
                                  end_row=row - 1, end_column=column)
                sheet.cell(row=start, column=column).alignment = Alignment(vertical="center")


# ---------------------------------------------------------------------------
# Blatt: Kompatibilitaet (lebende Matrix)
# ---------------------------------------------------------------------------
def sheet_matrix(book, parts) -> None:
    sheet = book.create_sheet("Kompatibilitaet")
    row = title(
        sheet, 1, "Was passt an welchen Slot",
        "Lebende Matrix: die Zellen rechnen aus Klasse und Bauart. Wer in den "
        "Spalten C bis F etwas aendert, sieht sofort, was daraus folgt.",
    )

    equipment = [p for p in parts if base_type(p["type"]) == "equipment"]
    bodies = [p for p in parts if base_type(p["type"]) == "body"]

    columns = []          # (Chassis-Name, Slotname, Regel)
    for body in bodies:
        rules = body.get("slot_rules", {})
        for slot in equip_slots(body):
            columns.append((body.get("name", body["id"]), slot, rules.get(slot)))

    head_chassis, head_slot, head_rule = row, row + 1, row + 2
    rank_row, cats_row = row + 3, row + 4
    header = row + 5
    first_data = header + 1

    labels = ["Code", "Teil", "Klasse", "Rang", "Bauart", "nur an Slot"]
    for index, label in enumerate(["Chassis", "Slot", "nimmt an", "Rang (Hilfe)",
                                   "Bauarten (Hilfe)"], start=0):
        cell = sheet.cell(row=row + index, column=6, value=label)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(horizontal="right")

    for offset, (chassis, slot, rule) in enumerate(columns):
        column = 7 + offset
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = 13

        cell = sheet.cell(row=head_chassis, column=column, value=chassis)
        cell.font = SUB_FONT
        cell.fill = SUB_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="bottom")

        cell = sheet.cell(row=head_slot, column=column, value=slot)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = Alignment(horizontal="center")

        cell = sheet.cell(row=head_rule, column=column, value=rule_text(rule))
        cell.font = Font(name=FONT, size=9, italic=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")

        limit = (rule or {}).get("max_class", "heavy")
        cell = sheet.cell(row=rank_row, column=column,
                          value=gen.MOUNT_CLASSES.index(limit) + 1)
        cell.font = NOTE_FONT
        cell.fill = HELPER_FILL
        cell.alignment = Alignment(horizontal="center")

        allowed = (rule or {}).get("categories") or list(gen.EQUIP_CATEGORIES)
        cell = sheet.cell(row=cats_row, column=column,
                          value=",".join(CATEGORY_LABEL[c] for c in allowed))
        cell.font = NOTE_FONT
        cell.fill = HELPER_FILL
        cell.alignment = Alignment(horizontal="center")

        cell = sheet.cell(row=header, column=column, value=slot.replace("equip_", ""))
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center")

    sheet.row_dimensions[head_chassis].height = 28
    sheet.row_dimensions[head_rule].height = 26
    header_row(sheet, header, labels)
    widths(sheet, {"A": 10, "B": 24, "C": 11, "D": 7, "E": 11, "F": 14})

    for index, part in enumerate(equipment):
        data_row = first_data + index
        # Volle Ankernamen, damit SEARCH() den Slotnamen der Spalte direkt
        # darin findet -- auch wenn ein Teil spaeter an mehrere Anker gebunden
        # wird und hier zwei Namen stehen.
        binding = ",".join(part.get("slots", [])) or "alle"
        values = [part.get("code", ""), part.get("name", part["id"]),
                  CLASS_LABEL[part.get("mount_class", "light")], None,
                  CATEGORY_LABEL[part.get("category", "weapon")], binding]
        for offset, value in enumerate(values, start=1):
            cell = sheet.cell(row=data_row, column=offset, value=value)
            cell.font = BODY_FONT if offset != 1 else CODE_FONT
            cell.border = BOX
        # Rang der Montageklasse -- als Formel, damit eine geaenderte Klasse
        # in Spalte C sofort durch die ganze Matrix laeuft.
        rank = sheet.cell(row=data_row, column=4,
                          value=f'=MATCH(C{data_row},$B$1:$D$1,0)')
        rank.font = NOTE_FONT
        rank.fill = HELPER_FILL
        rank.alignment = Alignment(horizontal="center")
        rank.border = BOX

        for offset in range(len(columns)):
            column = 7 + offset
            letter = get_column_letter(column)
            cell = sheet.cell(
                row=data_row, column=column,
                value=(f'=IF(AND($D{data_row}<={letter}${rank_row},'
                       f'ISNUMBER(SEARCH($E{data_row},{letter}${cats_row})),'
                       f'OR($F{data_row}="alle",'
                       f'ISNUMBER(SEARCH({letter}${head_slot},$F{data_row})))),'
                       f'"ja","–")'))
            cell.font = BODY_FONT
            cell.border = BOX
            cell.alignment = Alignment(horizontal="center")

    # Nachschlagereihe fuer MATCH: leicht/mittel/schwer in B1:D1.
    for offset, name in enumerate(("leicht", "mittel", "schwer")):
        cell = sheet.cell(row=1, column=2 + offset, value=name)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(horizontal="center")
    sheet.cell(row=1, column=5,
               value="← Rangfolge der Montageklassen, "
                     "Nachschlagereihe fuer Spalte D").font = NOTE_FONT

    sheet.freeze_panes = sheet.cell(row=first_data, column=7).coordinate


# ---------------------------------------------------------------------------
# Blatt: Silhouetten-Abstand
# ---------------------------------------------------------------------------
def sheet_silhouettes(book) -> None:
    sheet = book.create_sheet("Silhouetten")
    widths(sheet, {"A": 14, "B": 26, "C": 26, "D": 12, "E": 46})
    row = title(
        sheet, 1, "Silhouetten-Abstand",
        f"1.00 hiesse ununterscheidbar. Bei AUSRUESTUNG ist "
        f"{gen.SIL_ERROR:.2f} eine harte Grenze -- dort bricht der Generator "
        f"ab, weil der Spieler an der Waffe ihre Rolle abliest. Bei "
        f"RAHMENTEILEN gibt es KEINE Grenze und kein Urteil: das Mass "
        f"saettigt dort, jedes gedrungene Teil wird normiert zum selben "
        f"Klumpen. Brauchbar ist nur die Reihenfolge -- welches Paar liegt "
        f"am engsten. Ob die Formen wirklich verschieden sind, sagt das Auge.",
    )
    row = header_row(sheet, row, ["Typ", "Teil A", "Teil B", "Abstand", "Bewertung"])

    by_type: dict[str, list] = {}
    for spec, part in gen.all_parts():
        kind = base_type(part["type"])
        by_type.setdefault(kind, []).append(
            (part["code"], part.get("name", part["id"]), gen.silhouette(part["shapes"])))

    for part_type in TYPE_ORDER:
        entries = by_type.get(part_type, [])
        pairs = []
        for index, (code_a, name_a, sil_a) in enumerate(entries):
            for code_b, name_b, sil_b in entries[index + 1:]:
                pairs.append((gen.silhouette_distance(sil_a, sil_b),
                              f"{code_a} {name_a}", f"{code_b} {name_b}"))
        pairs.sort(reverse=True)
        strict = part_type == "equipment"
        for rank, (score, name_a, name_b) in enumerate(pairs[:3]):
            if not strict:
                # Rahmenteile bekommen KEIN Urteil. Das Mass saettigt dort --
                # zwei kompakte Teile decken sich als Flaeche, egal wie
                # verschieden sie gebaut sind. Was bleibt, ist die Reihenfolge.
                verdict = "engstes Paar dieses Typs" if rank == 0 else ""
            elif score >= gen.SIL_ERROR:
                verdict = "VERLETZT die Lesbarkeitsregel"
            elif score >= gen.SIL_ERROR - 0.08:
                verdict = "grenzwertig -- im Auge behalten"
            else:
                verdict = "unterscheidbar"
            for offset, value in enumerate(
                    [TYPE_LABEL[part_type], name_a, name_b, round(score, 2), verdict],
                    start=1):
                cell = sheet.cell(row=row, column=offset, value=value)
                cell.font = BODY_FONT
                cell.border = BOX
            sheet.cell(row=row, column=4).number_format = "0.00"
            sheet.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            row += 1

    row += 1
    cell = sheet.cell(row=row, column=1,
                      value="Pruefstein: die erste Belagerungskanone (ein Blaster "
                            "mit groesseren Zahlen) liegt bei "
                            f"{gen.silhouette_distance(gen.silhouette(gen.EQ_BLASTER), gen.silhouette(gen.EQ_CANNON_V1)):.2f}"
                            " -- genau der Fall, den die Regel verbietet.")
    cell.font = NOTE_FONT


# ---------------------------------------------------------------------------
# Selbstpruefung
# ---------------------------------------------------------------------------
# Ein winziger Formel-Rechner. Er kann genau die Funktionen, die diese Mappe
# benutzt -- mehr soll er nicht koennen. Sein Zweck ist, die geschriebene
# Formel selbst auszuwerten statt einer nachgebauten Zweitfassung: nur so
# faellt ein verrutschter Bezug auf.
_TOKEN = re.compile(
    r"""\s*("(?:[^"]|"")*"|"""
    r"""(?:'[^']+'|[A-Za-z][A-Za-z0-9_]*)!\$?[A-Z]+\$?\d+|"""
    r"""\$?[A-Z]+\$?\d+|[A-Za-z]+(?=\()|"""
    r"""<=|>=|<>|[(),=<>&:+\-*/]|\d+(?:\.\d+)?)""")


def _numbers(values) -> list[float]:
    return [v for v in values
            if isinstance(v, (int, float)) and not isinstance(v, bool)]


def _resolve(sheet, book, reference: str):
    """Trennt einen Blattbezug ab: 'Alle Teile'!$D$5 -> (Blatt, $D$5)."""
    if "!" not in reference:
        return sheet, reference
    name, reference = reference.split("!", 1)
    if book is None:
        raise SystemExit(f"Blattbezug {name}! ohne Mappe nicht aufloesbar")
    return book[name.strip("'")], reference


def _range(sheet, book, start: str, end: str) -> list:
    """Werte eines Zellbereichs, zeilenweise. Formeln werden mitgerechnet."""
    from openpyxl.utils import range_boundaries
    # Der Blattname steht in Excel nur am Anfang des Bereichs.
    sheet, start = _resolve(sheet, book, start)
    _, end = _resolve(sheet, book, end)
    c0, r0, c1, r1 = range_boundaries(f"{start}:{end}".replace("$", ""))
    out = []
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            value = sheet.cell(row=r, column=c).value
            out.append(evaluate(sheet, value, book)
                       if isinstance(value, str) else value)
    return out


def evaluate(sheet, formula: str, book=None):
    """Wertet eine Formel dieser Mappe aus."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    tokens, position = [], 1
    while position < len(formula):
        match = _TOKEN.match(formula, position)
        if not match:
            raise SystemExit(f"Formel nicht lesbar ab {formula[position:][:30]!r}")
        tokens.append(match.group(1))
        position = match.end()

    index = 0

    def peek():
        return tokens[index] if index < len(tokens) else None

    def take(expected=None):
        nonlocal index
        token = tokens[index]
        if expected and token != expected:
            raise SystemExit(f"Erwartet {expected!r}, gefunden {token!r}")
        index += 1
        return token

    def cell(reference: str):
        target, reference = _resolve(sheet, book, reference)
        column, row = "", ""
        for char in reference.replace("$", ""):
            (column := column + char) if char.isalpha() else (row := row + char)
        from openpyxl.utils import column_index_from_string
        value = target.cell(row=int(row),
                            column=column_index_from_string(column)).value
        # Verweist eine Formel auf eine andere Formel, wird die mit
        # ausgerechnet -- sonst vergliche man gegen einen Formeltext.
        return evaluate(target, value, book) if isinstance(value, str) else value

    def call(name: str):
        take("(")
        args = []
        while True:
            args.append(expression())
            if peek() == ",":
                take(",")
                continue
            take(")")
            break
        if name == "IF":
            return args[1] if args[0] else args[2]
        if name == "AND":
            return all(args)
        if name == "OR":
            return any(args)
        if name == "ISNUMBER":
            return isinstance(args[0], (int, float)) and not isinstance(args[0], bool)
        if name == "SEARCH":
            needle, haystack = str(args[0]).lower(), str(args[1]).lower()
            return haystack.index(needle) + 1 if needle in haystack else "#VALUE!"
        if name == "MATCH":
            return args[1].index(args[0]) + 1 if args[0] in args[1] else "#N/A"
        if name in ("MIN", "MAX", "AVERAGE", "SUM"):
            values = _numbers(args[0] if isinstance(args[0], list) else args)
            if not values:
                return "#DIV/0!" if name == "AVERAGE" else 0
            return {"MIN": min, "MAX": max, "SUM": sum,
                    "AVERAGE": lambda v: sum(v) / len(v)}[name](values)
        if name == "COUNTIF":
            return sum(1 for value in args[0] if value == args[1])
        if name == "ROUND":
            return round(args[0], int(args[1]))
        if name == "ABS":
            return abs(args[0])
        raise SystemExit(f"Funktion {name} kennt der Pruef-Rechner nicht")

    def atom():
        token = take()
        if token == "-":
            return -atom()
        if token == "(":
            value = expression()
            take(")")
            return value
        if token.startswith('"'):
            return token[1:-1].replace('""', '"')
        if token.replace(".", "").isdigit():
            return float(token) if "." in token else int(token)
        if token.isalpha():
            return call(token)
        if peek() == ":":                      # Bereich, z. B. $B$1:$D$1
            take(":")
            return _range(sheet, book, token, take())
        return cell(token)

    def number(value):
        # Eine leere Zelle ist in Excel eine Null, sobald damit gerechnet
        # wird. Das Blatt „Alle Teile“ laesst Nullen bewusst weg -- ohne
        # diese Regel rechnete der Pruefer dort etwas anderes als Excel.
        return 0 if value is None else value

    def product():
        left = number(atom())
        while peek() in ("*", "/"):
            operator = take()
            right = number(atom())
            left = left * right if operator == "*" else left / right
        return left

    def total():
        left = product()
        while peek() in ("+", "-"):
            operator = take()
            right = product()
            left = left + right if operator == "+" else left - right
        return left

    def expression():
        left = total()
        while peek() == "&":
            take("&")
            left = f"{left}{total()}"
        if peek() in ("<=", ">=", "<>", "=", "<", ">"):
            operator = take()
            right = expression()
            if isinstance(left, str) and "#" in str(left):
                return False
            return {"<=": lambda a, b: a <= b, ">=": lambda a, b: a >= b,
                    "=": lambda a, b: a == b, "<>": lambda a, b: a != b,
                    "<": lambda a, b: a < b, ">": lambda a, b: a > b}[operator](left, right)
        return left

    return expression()


def verify(path: pathlib.Path, parts: list[dict], layout: dict) -> None:
    """
    Rechnet die Formeln der Mappe nach -- ohne Excel.

    Eine Formel mit einem verrutschten Bezug faellt beim Ansehen nicht auf:
    sie liefert einfach eine falsche Antwort. Hier wird deshalb dieselbe Frage
    zweimal gestellt -- einmal an die Zellen der Mappe, einmal an die Regel
    selbst -- und beide Antworten muessen uebereinstimmen.
    """
    from openpyxl import load_workbook

    book = load_workbook(path)
    checked = _verify_matrix(book, parts)
    checked += _verify_threat(book, layout)
    checked += _verify_counts(book, parts)
    print(f"[ok] {checked} Formelzellen gegen die Regel geprueft, "
          f"kein Widerspruch")


def _verify_matrix(book, parts: list[dict]) -> int:
    """Die Kompatibilitaetsmatrix gegen fits_rule() im Generator."""
    sheet = book["Kompatibilitaet"]

    ranks = {sheet.cell(row=1, column=column).value: column - 1
             for column in (2, 3, 4)}
    if ranks != {"leicht": 1, "mittel": 2, "schwer": 3}:
        raise SystemExit(f"Nachschlagereihe B1:D1 stimmt nicht: {ranks}")

    head_slot = next(r for r in range(1, 20)
                     if sheet.cell(row=r, column=6).value == "Slot")
    first_data = head_slot + 5

    equipment = {p["code"]: p for p in parts
                 if base_type(p["type"]) == "equipment"}
    bodies = {}
    for part in parts:
        if base_type(part["type"]) == "body":
            for slot in equip_slots(part):
                bodies[(part.get("name", part["id"]), slot)] = \
                    part.get("slot_rules", {}).get(slot)

    columns = []
    column = 7
    while sheet.cell(row=head_slot, column=column).value:
        columns.append((column,
                        sheet.cell(row=head_slot - 1, column=column).value,
                        sheet.cell(row=head_slot, column=column).value))
        column += 1
    if len(columns) != len(bodies):
        raise SystemExit(f"{len(columns)} Slotspalten, aber {len(bodies)} Slots")

    checked = 0
    row = first_data
    while sheet.cell(row=row, column=1).value:
        code = sheet.cell(row=row, column=1).value
        part = equipment[code]
        for column, chassis, slot in columns:
            rule = bodies[(chassis, slot)]
            # Was die Regel sagt ...
            expected = (
                gen.fits_rule(part.get("mount_class", gen.DEFAULT_MOUNT_CLASS),
                              part.get("category", gen.DEFAULT_CATEGORY), rule)
                and (not part.get("slots") or slot in part["slots"])
            )
            # ... und was die FORMEL in der Zelle tatsaechlich rechnet.
            # Nicht was sie rechnen soll: eine nachgebaute Zweitfassung wuerde
            # denselben Denkfehler zweimal machen und ihn deshalb nie finden.
            actual = evaluate(sheet, sheet.cell(row=row, column=column).value)
            if actual != ("ja" if expected else "–"):
                raise SystemExit(
                    f"Matrix widerspricht der Regel: {code} an "
                    f"{chassis}/{slot} -- Regel sagt "
                    f"{'ja' if expected else 'nein'}, Formel rechnet {actual!r}")
            checked += 1
        row += 1
    return checked


def _verify_threat(book, layout: dict) -> int:
    """
    Die Threat-Spalten gegen threat_score() aus drome_build.gd.

    Geprueft wird zweierlei. Erstens: rechnet die Formel in der Zelle
    denselben Beitrag aus wie die Gewichtung des Spiels. Zweitens -- und das
    ist der eigentliche Punkt -- sind die Werte, die im Blatt gar nicht als
    Spalte stehen und in der Formel deshalb fehlen, fuer JEDES Teil dieses
    Typs wirklich null. Sonst unterschlaegt die Spalte etwas, ohne dass man
    es der Formel ansieht.
    """
    checked = 0
    for sheet_name, spec in layout.items():
        sheet = book[sheet_name]
        group, keys, first = spec["group"], spec["keys"], spec["first"]
        column = next(c for c in range(1, sheet.max_column + 1)
                      if sheet.cell(row=first - 1, column=c).value
                      == "Threat-Beitrag")

        for key in THREAT_WEIGHTS:
            if key in keys:
                continue
            for part in group:
                if stat(part, key) != 0:
                    raise SystemExit(
                        f"{sheet_name}: {part['code']} hat {key}="
                        f"{stat(part, key)}, aber die Threat-Formel dieses "
                        f"Blatts fuehrt {key} nicht mit")

        for offset, part in enumerate(group):
            row = first + offset
            if sheet.cell(row=row, column=1).value != part.get("code"):
                raise SystemExit(
                    f"{sheet_name}, Zeile {row}: erwartet {part.get('code')}, "
                    f"gefunden {sheet.cell(row=row, column=1).value!r}")
            actual = evaluate(sheet, sheet.cell(row=row, column=column).value)
            expected = threat_of(part, keys)
            if abs(actual - expected) > 1e-9:
                raise SystemExit(
                    f"{sheet_name}: Threat-Beitrag von {part['code']} rechnet "
                    f"{actual}, die Gewichtung ergibt {expected}")
            checked += 1

        if not spec["summary"]:
            continue

        # Und die Kennzahlen darunter: MIN/AVERAGE/MAX muessen genau die
        # Datenzeilen treffen, nicht eine zu viel oder zu wenig.
        summary = first + len(group) + 1
        values = [threat_of(p, keys) for p in group]
        for label, function in (("Minimum", min), ("Mittel", None), ("Maximum", max)):
            if sheet.cell(row=summary, column=1).value != label:
                raise SystemExit(
                    f"{sheet_name}: Kennzahl {label} steht nicht in Zeile {summary}")
            expected = (sum(values) / len(values) if function is None
                        else function(values))
            actual = evaluate(sheet, sheet.cell(row=summary, column=column).value)
            if abs(actual - expected) > 1e-9:
                raise SystemExit(
                    f"{sheet_name}: {label} ueber Threat rechnet {actual}, "
                    f"erwartet {expected}")
            checked += 1
            summary += 1
    return checked


def _verify_counts(book, parts: list[dict]) -> int:
    """Die COUNTIF-Bereiche der Legende muessen die Datenzeilen treffen."""
    legend = book["Legende"]
    checked = 0
    for row in range(1, legend.max_row + 1):
        formula = legend.cell(row=row, column=2).value
        if not isinstance(formula, str) or not formula.startswith("=COUNTIF"):
            continue
        label = legend.cell(row=row, column=1).value
        expected = sum(1 for p in parts if TYPE_LABEL[base_type(p["type"])] == label)
        actual = evaluate(legend, formula, book)
        if actual != expected:
            raise SystemExit(f"COUNTIF fuer {label} zaehlt {actual} statt {expected}")
        checked += 1
    return checked


# ---------------------------------------------------------------------------
def build(target: pathlib.Path) -> tuple[int, int]:
    parts, set_names = load_parts()
    book = Workbook()
    book.remove(book.active)

    # Reihenfolge der Blaetter ist die Reihenfolge der Erzeugung. Die Legende
    # entsteht zuletzt, weil sie die Zeilenspanne von „Alle Teile“ braucht,
    # und wird danach nach vorn geschoben.
    parts_range = sheet_all(book, parts, set_names)

    # Blatt -> was dort steht, mit welcher Threat-Formel, ab welcher Zeile.
    # Die Selbstpruefung liest daraus, welche Formel wo stehen muss.
    layout: dict[str, dict] = {
        "Alle Teile": {"group": parts, "keys": list(THREAT_WEIGHTS),
                       "first": parts_range[0], "summary": False},
    }

    def add(name, heading, note, part_type, columns, threat_keys, categories=None):
        group = [p for p in parts
                 if base_type(p["type"]) == part_type
                 and (categories is None or p.get("category") in categories)]
        columns = columns + [threat_column(threat_keys)] + reading_columns()
        sheet_type(book, name, heading, note, group, columns)
        # Titel (Zeile 1) + Untertitel (2) + Leerzeile (3) + Kopfzeile (4)
        layout[name] = {"group": group, "keys": threat_keys,
                        "first": 5, "summary": True}

    add("Koerperteile",
        "Koerperteile -- Chassis",
        "Das Chassis traegt alles andere: es bringt den Grossteil der HP, "
        "setzt die Traglast und bestimmt ueber seine equip_*-Anker, wie viele "
        "Ausruestungsslots der DROME ueberhaupt hat.",
        "body",
        head_columns(set_names) + [
            stat_column("hp"),
            stat_column("def"),
            stat_column("spd"),
            stat_column("mov"),
            stat_column("weight_capacity", 10),
            stat_column("en_max", 9),
            Column("slots", "Slots", 7,
                   lambda p: len(equip_slots(p)), fmt="0", summary=True,
                   note="Zahl der equip_*-Anker. Ein Chassis bekommt einen "
                        "weiteren Slot, indem es einen weiteren Anker in "
                        "seine JSON schreibt -- kein Sonderfall im Code."),
            Column("slot_rules", "Slots und ihre Regel", 34,
                   lambda p: "\n".join(
                       f"{s.replace('equip_', '')}: "
                       f"{rule_text(p.get('slot_rules', {}).get(s))}"
                       for s in equip_slots(p)), wrap=True),
        ],
        ["hp", "def", "spd", "mov", "en_max"])

    add("Kopfteile",
        "Kopfteile -- Sensorik",
        "Der Kopf entscheidet, was der DROME sieht und wie genau er trifft. "
        "Nur der Molok-Bunkerkopf bringt zusaetzlich Struktur mit.",
        "head",
        head_columns(set_names) + [
            stat_column("atk"),
            stat_column("def"),
            stat_column("hp"),
            stat_column("spd"),
            stat_column("en_max", 9),
            stat_column("en_regen", 10),
            stat_column("weight"),
            stat_column("power_draw", 12),
            flag_column("grants_ignore_haze", 13),
        ],
        ["hp", "atk", "def", "spd", "en_max"])

    add("Fussteile",
        "Fussteile -- Antrieb",
        "Der Antrieb ist die eigentliche Kaufentscheidung im Gelaende: nicht "
        "MOV allein, sondern welche Felder ueberhaupt betretbar sind. Die "
        "Flags rechts sind das, was ein Antrieb kann und ein anderer nicht.",
        "feet",
        head_columns(set_names) + [
            stat_column("mov"),
            stat_column("spd"),
            stat_column("def"),
            stat_column("weight"),
            stat_column("power_draw", 12),
            flag_column("can_pass_units", 12),
            flag_column("can_enter_steps", 9),
            flag_column("ignores_drift", 10),
            stat_column("drift_modifier", 9),
            flag_column("step_cost_reduced", 12),
        ],
        ["def", "spd", "mov"])

    add("Kerne",
        "Kerne -- Energie",
        "Der Kern setzt zwei Grenzen: EN max deckelt, wie oft Aktionen mit "
        "EN-Kosten gehen, und die Energieabgabe deckelt, wie viel Geraet "
        "ueberhaupt angebaut werden darf. Ein Kern ist damit weniger eine "
        "Waffe als ein Budget.",
        "core",
        head_columns(set_names) + [
            stat_column("en_max", 10),
            stat_column("en_regen", 10),
            stat_column("power_output", 13),
            stat_column("atk"),
            stat_column("weight"),
        ],
        ["atk", "en_max"])

    equipment_columns = head_columns(set_names) + [
        Column("mount_class", "Montageklasse", 12,
               lambda p: CLASS_LABEL.get(p.get("mount_class", "light")),
               note="Was die HALTERUNG aushalten muss -- Masse, Rueckstoss, "
                    "Hebel. Nicht das Gewicht allein. Welcher Slot welche "
                    "Klasse nimmt, steht auf dem Blatt „Chassis & Slots“."),
        stat_column("weight"),
        stat_column("power_draw", 12),
        stat_column("atk"),
        stat_column("def"),
        Column("act_name", "Aktion", 20,
               lambda p: (action(p) or {}).get("display_name", "—"),
               note="Passive Teile gewaehren keine Aktion."),
        Column("act_targeting", "Zielung", 16,
               lambda p: TARGETING_LABEL.get(
                   (action(p) or {}).get("targeting"), "")),
        action_column("power"),
        action_column("range_tiles", 11),
        action_column("aoe_radius", 8),
        action_column("en_cost", 10),
        Column("act_los", "Sichtlinie", 10,
               lambda p: ("ja" if (action(p) or {}).get("requires_line_of_sight")
                          else "–") if action(p) else None,
               note=ACTION_GLOSSARY["requires_line_of_sight"][1]),
        action_column("push_tiles", 8, summary=False),
    ]

    add("Ausruestung Waffen",
        "Ausruestung -- Waffen",
        "Bauart „Waffe“. Ohne mindestens eine davon ist ein Aufbau ungueltig. "
        "Die Spalte „Power je EN“ vergleicht, was ein Schuss kostet: eine "
        "starke Waffe mit hohen EN-Kosten ist auf einem schwachen Kern "
        "weniger wert als eine schwaechere, die nichts kostet.",
        "equipment",
        equipment_columns + [
            Column("per_en", "Power je EN", 11,
                   formula=lambda L, r: (
                       f'=IF({L["act_en_cost"]}{r}=0,"kostenlos",'
                       f'ROUND({L["act_power"]}{r}/{L["act_en_cost"]}{r},1))'),
                   note="Schaden je Energiepunkt. „kostenlos“ heisst: die "
                        "Aktion kostet gar keine Energie und laesst sich "
                        "jeden Zug einsetzen."),
            Column("per_weight", "Power je Gewicht", 12,
                   formula=lambda L, r: (
                       f'=ROUND({L["act_power"]}{r}/{L["weight"]}{r},1)'),
                   fmt="0.0", summary=True,
                   note="Schaden je Gewichtspunkt -- die Waehrung, in der die "
                        "Traglast des Chassis bezahlt wird."),
            Column("best_weapon", "Threat als staerkste Waffe", 14,
                   formula=lambda L, r: (
                       f'={L["act_power"]}{r}*{THREAT_BEST_WEAPON:g}'),
                   fmt="0.0", summary=True,
                   note="Zusatzbeitrag zum Threat Score -- aber nur fuer die "
                        "staerkste Waffe eines Aufbaus, nicht fuer jede. "
                        "Deshalb steht der Wert neben dem Threat-Beitrag und "
                        "nicht darin."),
        ],
        ["atk", "def"], categories=["weapon"])

    # Schild und Support teilen sich ein Blatt -- „Sonstiges“ ist genau das.
    # Die Bauart wird deshalb hier zur eigenen Spalte: auf dem Waffenblatt
    # steht in ihr sonst 7-mal dasselbe Wort.
    add("Ausruestung Sonstiges",
        "Ausruestung -- Schild und Support",
        "Alles, was keine Waffe ist. Der Deflektor ist rein passiv (DEF, "
        "keine Aktion); die Support-Module gewaehren stattdessen eine "
        "Faehigkeit. Negative Power heisst Heilung -- ein Vorzeichen statt "
        "zweier Felder.",
        "equipment",
        equipment_columns[:4] + [
            Column("category", "Bauart", 10,
                   lambda p: CATEGORY_LABEL.get(p.get("category"), "")),
        ] + equipment_columns[4:] + [
            Column("effect", "Wirkung", 16,
                   formula=lambda L, r: (
                       f'=IF({L["act_name"]}{r}="—","passiv",'
                       f'IF({L["act_power"]}{r}<0,'
                       f'"Heilung "&ABS({L["act_power"]}{r}),'
                       f'IF({L["act_power"]}{r}=0,"kein Schaden",'
                       f'"Schaden "&{L["act_power"]}{r})))'),
                   note="Liest die Spalte „Power“: positiv ist Schaden, "
                        "negativ Heilung, 0 heisst reine Verschiebung oder "
                        "Status."),
        ],
        ["atk", "def"], categories=["shield", "support"])

    sheet_chassis(book, parts, set_names)
    sheet_matrix(book, parts)
    sheet_silhouettes(book)

    type_letter = get_column_letter(4)      # Spalte „Typ“ auf „Alle Teile“
    sheet_legend(book, parts_range, type_letter)
    book.move_sheet("Legende", offset=-(len(book.sheetnames) - 1))

    target.parent.mkdir(parents=True, exist_ok=True)
    book.save(target)
    verify(target, parts, layout)
    return len(parts), len(book.sheetnames)


def main() -> None:
    target = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else \
        ROOT / "docs" / "cyberdrome_teile.xlsx"
    count, sheets = build(target)
    print(f"[ok] {target}  ({count} Teile, {sheets} Blaetter)")


if __name__ == "__main__":
    main()
