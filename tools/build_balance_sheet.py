#!/usr/bin/env python3
"""
build_balance_sheet.py -- Ausruestungs-Statistik-Liste als Excel-Mappe.

    python3 tools/build_balance_sheet.py [ziel.xlsx]

Liest die kuratierte Balancing-Quelle data/balancing/ausruestung_stats.json und
schreibt daraus eine gut lesbare Arbeitsmappe (Vorgabe: docs/ausruestung_stats.xlsx).

Zweck
-----
Das ist die schlanke BALANCING-Sicht auf das Klassenmodell (GAME_DESIGN §9/§12):
die Huelle als EIN Teil, wenige aussagekraeftige Werte je Kategorie. Nicht zu
verwechseln mit docs/cyberdrome_teile.xlsx -- die kommt aus den Teil-JSONs und
zeigt den vollen Engine-Wertesatz. Diese Mappe hier ist zum Anfassen und Tunen.

Eine Quelle der Wahrheit
------------------------
Alle Zahlen kommen aus der JSON. Wer einen Wert aendert, aendert ihn dort (oder
in dieser Mappe und schickt sie zurueck) und laesst die Mappe neu erzeugen --
eine von Hand doppelt gefuehrte Liste laeuft sonst auseinander.
"""

from __future__ import annotations

import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
SOURCE = ROOT / "data" / "balancing" / "ausruestung_stats.json"
DEFAULT_OUT = ROOT / "docs" / "ausruestung_stats.xlsx"

# --- Optik ------------------------------------------------------------------
# Eine ruhige Palette: dunkle Kopfzeile je Kategorie, helle Zebrastreifen, die
# Werte-Spalten leicht hervorgehoben, damit klar ist, was hier getunt wird.
HEAD_FILL = PatternFill("solid", fgColor="1F2A44")   # tiefes Nachtblau
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="1F2A44", size=15)
SUB_FONT = Font(italic=True, color="55607A", size=10)
STAT_FILL = PatternFill("solid", fgColor="EAF1FB")   # Werte-Spalten
ZEBRA_FILL = PatternFill("solid", fgColor="F5F7FB")
STAT_FONT = Font(bold=True, color="112233", size=11)
WRAP = Alignment(vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D5DBE7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load() -> dict:
    with SOURCE.open(encoding="utf-8") as fh:
        return json.load(fh)


def _title(sheet, text: str, note: str = "") -> int:
    """Titelzeile plus optionaler Unterzeile. Gibt die naechste freie Zeile."""
    cell = sheet.cell(row=1, column=1, value=text)
    cell.font = TITLE_FONT
    row = 2
    if note:
        cell = sheet.cell(row=row, column=1, value=note)
        cell.font = SUB_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 28
        row += 1
    return row + 1


def _table(sheet, start_row: int, columns: list[dict], rows: list[dict],
           stat_keys: set[str]) -> int:
    """
    Schreibt eine Tabelle. ``columns`` ist eine Liste von {key, header, width,
    align}. ``stat_keys`` markiert die getunten Werte-Spalten (Hervorhebung).
    """
    head = start_row
    for offset, col in enumerate(columns, start=1):
        cell = sheet.cell(row=head, column=offset, value=col["header"])
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER if col.get("align") == "center" else LEFT
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(offset)].width = col["width"]
    sheet.row_dimensions[head].height = 26

    row = head + 1
    for r_index, item in enumerate(rows):
        for offset, col in enumerate(columns, start=1):
            value = item.get(col["key"], "")
            cell = sheet.cell(row=row, column=offset, value=value)
            cell.border = BORDER
            is_stat = col["key"] in stat_keys
            if is_stat:
                cell.fill = STAT_FILL
                cell.font = STAT_FONT
                cell.alignment = CENTER
            else:
                if r_index % 2 == 1:
                    cell.fill = ZEBRA_FILL
                cell.alignment = CENTER if col.get("align") == "center" else LEFT
        sheet.row_dimensions[row].height = col_row_height(item, columns)
        row += 1

    last = row - 1
    sheet.freeze_panes = sheet.cell(row=head + 1, column=1).coordinate
    sheet.auto_filter.ref = f"A{head}:{get_column_letter(len(columns))}{last}"
    return row + 1


def col_row_height(item: dict, columns: list[dict]) -> int:
    """Etwas hoehere Zeilen, wenn eine lange Text-Spalte umbrechen muss."""
    longest = 0
    for col in columns:
        if col.get("align") == "center":
            continue
        text = str(item.get(col["key"], ""))
        longest = max(longest, len(text) / max(col["width"], 1))
    return 18 + int(longest) * 14


def sheet_uebersicht(book: Workbook, data: dict) -> None:
    sheet = book.create_sheet("Uebersicht")
    sheet.sheet_view.showGridLines = False
    row = _title(sheet, "Ausruestungs-Statistik-Liste",
                 "Schlankes Balancing-Modell fuer den Reframe (GAME_DESIGN §9/§12). "
                 "Zahlen sind Vorschlaege -- zum Anpassen gedacht.")
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 70

    # Modell-Tabelle
    cell = sheet.cell(row=row, column=1, value="Wieviele Werte je Kategorie")
    cell.font = Font(bold=True, color="1F2A44", size=12)
    row += 1
    model_cols = [("Kategorie", 22), ("Werte", 16), ("Passive?", 70)]
    for offset, (header, _) in enumerate(model_cols, start=1):
        c = sheet.cell(row=row, column=offset, value=header)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    labels = {"huelle": "Huelle / Chassis", "waffe": "Waffe",
              "kern": "Kern", "gadget": "Gadget"}
    for key, spec in data["modell"].items():
        passive = ("ja -- genau eine Signatur-Passive je Huelle"
                   if spec.get("hat_passive") else "nein")
        values = (labels[key], spec["anzahl_stats"], passive)
        for offset, value in enumerate(values, start=1):
            c = sheet.cell(row=row, column=offset, value=value)
            c.border = BORDER
            c.alignment = CENTER if offset == 2 else LEFT
        row += 1
    row += 1

    # Glossar
    cell = sheet.cell(row=row, column=1, value="Was die Werte bedeuten")
    cell.font = Font(bold=True, color="1F2A44", size=12)
    row += 1
    for header in ("Wert", "Engine-Feld", "Beschreibung"):
        c = sheet.cell(row=row, column=("Wert", "Engine-Feld", "Beschreibung").index(header) + 1,
                       value=header)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    for spec in data["stat_glossar"].values():
        for offset, value in enumerate((spec["label"], spec["engine"],
                                        spec["beschreibung"]), start=1):
            c = sheet.cell(row=row, column=offset, value=value)
            c.border = BORDER
            c.alignment = LEFT
        row += 1
    row += 1

    # Hinweise aus dem JSON-Kommentar (die offenen Fragen sichtbar machen)
    cell = sheet.cell(row=row, column=1, value="Hinweise")
    cell.font = Font(bold=True, color="1F2A44", size=12)
    row += 1
    for line in data.get("_kommentar", []):
        c = sheet.cell(row=row, column=1, value=line)
        c.font = SUB_FONT if line else SUB_FONT
        c.alignment = LEFT
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1


def build(out_path: pathlib.Path) -> None:
    data = load()
    book = Workbook()
    book.remove(book.active)  # die leere Vorgabe-Seite

    sheet_uebersicht(book, data)

    # --- Huelle / Chassis --------------------------------------------------
    sheet = book.create_sheet("Huelle")
    sheet.sheet_view.showGridLines = False
    start = _title(sheet, "Huelle / Chassis -- 3 Werte + Signatur-Passive",
                   "Der ganze Rahmen als ein Teil (§9c). Passive: genau eine je Klasse.")
    _table(sheet, start,
           [
               {"key": "code", "header": "Code", "width": 10},
               {"key": "name", "header": "Name", "width": 18},
               {"key": "rolle", "header": "Rolle", "width": 34},
               {"key": "integritaet", "header": "Integritaet", "width": 12, "align": "center"},
               {"key": "panzerung", "header": "Panzerung", "width": 11, "align": "center"},
               {"key": "tempo", "header": "Tempo", "width": 9, "align": "center"},
               {"key": "passive_name", "header": "Passive", "width": 16},
               {"key": "passive_wirkung", "header": "Passive -- Wirkung", "width": 52},
               {"key": "notizen", "header": "Notizen", "width": 40},
           ],
           [
               {
                   **{k: h[k] for k in ("code", "name", "rolle", "integritaet",
                                        "panzerung", "tempo", "notizen")},
                   "passive_name": h["passive"]["name"],
                   "passive_wirkung": h["passive"]["wirkung"],
               }
               for h in data["huellen"]
           ],
           stat_keys={"integritaet", "panzerung", "tempo"})

    # --- Waffen ------------------------------------------------------------
    sheet = book.create_sheet("Waffen")
    sheet.sheet_view.showGridLines = False
    start = _title(sheet, "Waffen -- 2 Werte",
                   "Schaden und Reichweite -- die zwei Zahlen, die eine Waffe im "
                   "Tactics ausmachen (§3d).")
    _table(sheet, start,
           [
               {"key": "code", "header": "Code", "width": 10},
               {"key": "name", "header": "Name", "width": 20},
               {"key": "art", "header": "Art", "width": 26},
               {"key": "fuer", "header": "Chassis", "width": 10},
               {"key": "schaden", "header": "Schaden", "width": 10, "align": "center"},
               {"key": "reichweite", "header": "Reichweite", "width": 11, "align": "center"},
               {"key": "notizen", "header": "Notizen", "width": 50},
           ],
           data["waffen"],
           stat_keys={"schaden", "reichweite"})

    # --- Kerne -------------------------------------------------------------
    sheet = book.create_sheet("Kerne")
    sheet.sheet_view.showGridLines = False
    start = _title(sheet, "Kerne -- 2 Werte",
                   "Energie (Tank) und Regeneration. Die aktive Faehigkeit haengt "
                   "auch am Kern (§13), ist aber kein Stat.")
    _table(sheet, start,
           [
               {"key": "code", "header": "Code", "width": 10},
               {"key": "name", "header": "Name", "width": 22},
               {"key": "typ", "header": "Typ", "width": 22},
               {"key": "energie", "header": "Energie", "width": 10, "align": "center"},
               {"key": "regeneration", "header": "Regeneration", "width": 13, "align": "center"},
               {"key": "notizen", "header": "Notizen", "width": 50},
           ],
           data["kerne"],
           stat_keys={"energie", "regeneration"})

    # --- Gadgets -----------------------------------------------------------
    sheet = book.create_sheet("Gadgets")
    sheet.sheet_view.showGridLines = False
    start = _title(sheet, "Gadgets -- 1 Wert",
                   "Support und Schild. Je Teil GENAU EIN Wert -- aber je Teil ein "
                   "anderer (Schild, Heilung, Sog, Aggro).")
    _table(sheet, start,
           [
               {"key": "code", "header": "Code", "width": 10},
               {"key": "name", "header": "Name", "width": 20},
               {"key": "typ", "header": "Typ", "width": 12},
               {"key": "stat", "header": "Wirk-Wert", "width": 14},
               {"key": "wert", "header": "Wert", "width": 9, "align": "center"},
               {"key": "notizen", "header": "Notizen", "width": 54},
           ],
           data["gadgets"],
           stat_keys={"wert"})

    out_path.parent.mkdir(parents=True, exist_ok=True)
    book.save(out_path)
    print(f"[balance] geschrieben: {out_path.relative_to(ROOT)}")


def main(argv: list[str]) -> int:
    out = pathlib.Path(argv[1]).resolve() if len(argv) > 1 else DEFAULT_OUT
    if not SOURCE.exists():
        raise SystemExit(f"Quelle fehlt: {SOURCE}")
    build(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
